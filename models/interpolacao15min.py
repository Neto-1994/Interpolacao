from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image

# Busca de dados no banco
try:
    class Interpolar_15min():
        def calcular(self, data1, data2, data3, nomearquivo, nomesalvar):
            # Carregar arquivo excel existente
            wb = load_workbook(nomearquivo + ".xlsx")
            # Lista todos os nomes das planilhas no arquivo
            nomes_planilhas = wb.sheetnames
            # Carrega a planilha escolhida
            ws1 = wb[nomes_planilhas[0]]
            ws2 = wb[nomes_planilhas[1]]
            # Loop while para verificar a condição
            while data3 <= data2:
                linha1 = 0
                linha2 = 0
                # Percorrer as linhas da planilha1 até encontrar o valor
                for row in ws1.iter_rows(min_row=1, max_col=1, max_row=ws1.max_row, values_only=True):
                    linha1 += 1
                    if row[0] == data3:
                        # Obter o valor da coluna 2 na linha correspondente
                        x1 = ws1.cell(row=linha1 - 2, column=2).value
                        x2 = ws1.cell(row=linha1 - 1, column=2).value
                        x3 = ws1.cell(row=linha1, column=2).value
                        break
                # Converter a string da data para um objeto datetime
                data_i = datetime.strptime(data3, "%Y-%m-%d %H:%M:%S")
                # Formatar horário para a planilha2
                nova_data_hora = data_i - timedelta(minutes=15)
                # Formatar a nova data e hora de volta para string
                valor2 = nova_data_hora.strftime("%Y-%m-%d %H:%M:%S")
                # Percorrer as linhas da planilha2 até encontrar o valor
                for row in ws2.iter_rows(min_row=1, max_col=1, max_row=ws2.max_row, values_only=True):
                    linha2 += 1
                    if row[0] == valor2:
                        # Obter o valor da coluna 2 na linha correspondente
                        y1 = ws2.cell(row=linha2 - 1, column=2).value
                        y2 = ws2.cell(row=linha2, column=2).value
                        break
                # Validações dos valores
                if x1 == x2 and x2 == x3:
                    y3 = y2
                    # print("Passou no 1°..")
                else:
                    if x2 == x3:
                        y3 = y2
                        # print("Passou no 2°..")
                    else:
                        if x1 == x2:
                            # Parametro para descontar as linhas onde buscou valores
                            l = 3
                            while x1 == x2:
                                x1 = ws1.cell(row=linha1 - l, column=2).value
                                # print(f"Rebuscando x1: {x1}")
                                l += 1
                        if y1 == y2:
                            # Parametro para descontar as linhas onde buscou valores
                            l = 2
                            while y1 == y2:
                                y1 = ws2.cell(row=linha2 - l, column=2).value
                                # print(f"Rebuscando y1: {y1}")
                                l += 1
                        # print(f"Valor de x1: {x1}\n")
                        # print(f"Valor de x2: {x2}\n")
                        # print(f"Valor de x3: {x3}\n")
                        # print(f"Valor de y1: {y1}\n")
                        # print(f"Valor de y2: {y2}\n")
                        y3 = round(y1 + ((x3 - x1) * (y2 - y1)) / (x2 - x1), 2)
                        # print("Passou no 3°..")
                # Formatar a nova data e hora de volta para string
                data_salvar = data_i.strftime("%Y-%m-%d %H:%M:%S")
                # Inserir dados na planilha
                ws2.append({"A": data_salvar, "B": y3})
                # print(f"Data: {data_salvar}")
                # print(f"Valor de y3: {y3}\n")
                # print("-----------------------")
                # Incrementa 15 minutos a cada iteração
                data_i += timedelta(minutes=15)
                # Formatar a nova data e hora de volta para string
                data3 = data_i.strftime("%Y-%m-%d %H:%M:%S")
            # Leitura de parâmetros do arquivo
            linhas = ws2.max_row + 1
            colunas = ws2.max_column + 1
            # Formatar dados da planilha
            for i in range(1, linhas):
                for j in range(1, colunas):
                    ws2.cell(i, j).font = Font(name="Calibri",
                                                    size=12)
#                                             bold = False,
#                                             italic = False,
#                   planilha.cell(i, j).border = Border(left=Side(border_style="thin",
#                                                            color='FF000000'),
#                                                  right=Side(border_style="thin",
#                                                             color='FF000000'),
#                                                  top=Side(border_style="thin",
#                                                           color='FF000000'),
#                                                  bottom=Side(border_style="thin",
#                                                              color='FF000000'))
#                                                 diagonal=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 diagonal_direction=0,
#                                                 outline=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 vertical=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 horizontal=Side(border_style=None,
#                                                 color='FF000000'))

                    ws2.cell(i, j).alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws2.cell(i, j).number_format = '0.00'
            # Apresentacao dos dataframes no terminal
#            displayhook(df)
            # Feche o arquivo Excel
            wb.close()
            # Exportar dataframes como arquivo xlsx
            wb.save(nomesalvar + ".xlsx")  # Gerar arquivo pelo openpyxl
except OSError as e:
    print("Erro: ", e)
