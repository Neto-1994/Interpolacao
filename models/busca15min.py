from sys import displayhook
import pandas
import Conexao
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image

# Busca de dados no banco
try:
    class Busca15min():
        def buscar(self, data1, data2, codigo1, nomeSalvar):
            consulta_sql = "SELECT e.Nome_Estacao AS Estacao, m.HoraLocal AS Horario, m.SPressao AS SPressao FROM medicoes m JOIN estacoes e ON m.Codigo_Sec = e.Codigo_Sec WHERE m.Codigo_Sec = %s AND HoraLocal BETWEEN %s AND %s;"
            cursor = Conexao.obter_conexao().cursor()
            cursor.execute(consulta_sql, (codigo1, data1, data2))
            Dados = cursor.fetchall()
            # Encerrar conexao com o banco de dados
            try:
                cursor.close()
            except:
                print("\nErro ao fechar conexão MySQL..")
            # Gerar dataframe com os dados
            df = pandas.DataFrame(
                Dados, columns=["Estacao", "Horario", "Nivel"])
            # Obter o nome da estação da primeira linha do DataFrame
            nome = df.loc[0, "Estacao"]
            # Criar um novo arquivo excel
            arquivo = Workbook()
            # Adiciona uma planilha ao arquivo
            planilha = arquivo.active
            # Adiciona um nome à planilha
            planilha.title = nome
            # Modificar o DataFrame original eliminando coluna
            df.drop("Estacao", axis=1, inplace=True)
            # Formatacao da data
            df["Horario"] = pandas.to_datetime(df.Horario)
            # Ano com Y maiúsculo, saída com 4 dígitos / Ano com y minúsculo, saída com 2 dígitos
            df["Horario"] = df["Horario"].dt.strftime("%Y-%m-%d %H:%M:%S")
            # Transformar dataframe em datarows (linhas de dados)
            dr = dataframe_to_rows(df, index=False, header=True)
            # Inserir dados na planilha
            for r in dr:
                planilha.append(r)
            # Inserir imagens na planilha
#            img1 = Image("C:/Users/Jair/Pictures/Acqua.png")
#            img2 = Image("C:/Users/Jair/Pictures/Lundin.png")
#            ws.add_image(img1, "A1")
#            ws.add_image(img2, "D1")
            # Leitura de parâmetros do arquivo
            linhas = planilha.max_row + 1
            colunas = planilha.max_column + 1
            # Formatar dados da planilha
            for i in range(1, linhas):
                for j in range(1, colunas):
                    planilha.cell(i, j).font = Font(name="Calibri",
                                                    size=12)
#                                             bold = False,
#                                             italic = False,
#                   planilha.cell(i, j).border = Border(left=Side(border_style="thin",
#                                                            color='FF000000'),
#                                                  right=Side(border_style="thin",
#                                                             color='FF000000'),
#                                                  top=Side(border_style="thin",
#                                                           color='FF000000'),
#                                                  bottom=Side(border_style="thin",
#                                                              color='FF000000'))
#                                                 diagonal=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 diagonal_direction=0,
#                                                 outline=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 vertical=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 horizontal=Side(border_style=None,
#                                                 color='FF000000'))
                    planilha.cell(i, j).alignment = Alignment(
                        horizontal='center', vertical='center')
                    planilha.cell(i, j).number_format = '0.00'
            # Apresentacao dos dataframes no terminal
#           displayhook(df)
            # Exportar dataframes como arquivo xlsx
#           df.to_excel("Teste Salvamento.xlsx", index= False) # Gerar arquivo pelo pandas
            arquivo.save(nomeSalvar + ".xlsx")  # Gerar arquivo pelo openpyxl
except OSError as e:
    print("Erro: ", e)
